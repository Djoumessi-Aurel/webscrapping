import pandas as pd
from openpyxl import load_workbook
import re

def split_value(value: str):
    regex = r'Nom du membre(.*)Fonction du membre(.*)Connecté(.*)'
    match = re.match(regex, str(value))
    if match:
        return {"nom": match.group(1).strip(), "fonction": match.group(2).strip(), "connecte": match.group(3).strip()}
    else:
        return {"nom": None, "fonction": None, "connecte": None}


def filter_value(value: str):
    regex = r'Nom du membre(.*)Fonction du membre(.*)Connecté(.*)'
    match = re.match(regex, str(value))
    if match:
        return True
    else:
        return False
    
fichier = "LinkedIn.xlsx" # Chemin du fichier à traîter
    
# TRAITEMENT 1: EXTRACTION DU NOM, DE LA FONCTION ET DU MOMENT DE CONNECTION #
# Lire les données en ignorant les images
data = pd.read_excel(fichier, engine='openpyxl')

data = data[data["N°"]!="Message"]
data = data[~data["N°"].isna()]
data["filtre"] = data["N°"].apply(filter_value)
data = data[data["filtre"]]
data["dict"] = data["N°"].apply(split_value)

data["Nom"] = data["dict"].apply(lambda x: x["nom"])
data["Fonction"] = data["dict"].apply(lambda x: x["fonction"])
data["Connecté"] = data["dict"].apply(lambda x: x["connecte"])

data = data[["Nom", "Fonction", "Connecté"]].reset_index().drop(["index"], axis="columns")


# TRAITEMENT 2: EXTRACTION DU LIEN DE LA PAGE LINKEDIN ET ENREGISTREMENT DU FICHIER
# Charger le fichier Excel
wb = load_workbook(fichier)
ws = wb.active  # ou wb['nom_de_votre_feuille'] si vous connaissez le nom

# Pour stocker les données
donnees = []

# Ajustez le range si nécessaire selon votre fichier
for row in ws.iter_rows(min_row=2):  # min_row=2 pour ignorer l'en-tête
    cell = row[0]  # Première colonne (N°)
    
    texte = cell.value  # Le texte visible
    if not filter_value(texte):
        continue
    nom = split_value(texte)["nom"]

    url = None    
    if cell.hyperlink:
        # Récupérer l'URL du lien hypertexte
        url = cell.hyperlink.target
        
    donnees.append({'Nom': nom, 'Lien': url})

# Créer un DataFrame
df = pd.DataFrame(donnees)

# Joindre les deux dataframes et enregistrer le résultat
# data = data.join(df, on="Nom", how="inner")
data["Lien"] = df["Lien"]
data.to_excel("Processed_data.xlsx", index=False)
